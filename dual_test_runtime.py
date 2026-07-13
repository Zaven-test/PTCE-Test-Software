"""
dual_test_runtime.py

Runtime for a single alternating current-flow sequence shared between two
test setups:
- HV Connectors test (line0 + line1 contactors, staggered close)
- CCA Fuses test (line2 contactor)

Only one setup ever draws current at a time. The runtime cycles between
them automatically and indefinitely: flow on the active test, a grace
period with current off, then contactors flip and the other test flows.
Persistent state lets the script pause/resume and recover after restart.

Continuous shunt-current logging to Influx/CSV/XLSX lives in the separate
logger.py process (run it alongside this script) so logging keeps running
across restarts of this file. This script still owns the TC thermocouple
reads and the over-temp safety trip, since those gate its own pause logic,
and it publishes its PSU current/voltage readback to LIVE_READINGS_FILE for
logger.py to pick up.
"""

import csv
import json
import math
import os
import signal
import threading
import time
import traceback
from datetime import datetime

# ---------------------------------------------------------------------------
# Editable test configuration
# ---------------------------------------------------------------------------
SIMULATE = False
PSU_VISA_ADDRESS = 'USB0::0x2EC7::0x3900::805240031807340017::INSTR'
PSU_VOLTAGE_LIMIT = 10.0
RAMP_STEP_A = 10.0
RAMP_STEP_DELAY_S = 0.25
SHUNT_RESISTANCE_OHM = 50e-6
MAX_VOLTAGE_V = 10.0
LOG_INTERVAL = 10
STATE_SAVE_INTERVAL_S = 5
STATE_FILE = 'dual_test_state.json'
HISTORY_FILE = 'dual_test_history.log'
LOG_FILE_XLSX = 'dual_test_log.xlsx'
LOG_FILE_CSV_TEMPLATE = 'dual_test_log_{test}.csv'
# PSU current/voltage readback, refreshed every LOG_INTERVAL, so logger.py can
# include it in its rows without opening its own connection to the PSU.
LIVE_READINGS_FILE = 'dual_test_live_readings.json'

AUTO_START = False
GRACE_PERIOD_S = 15

# Contactors on the same 12V supply must not close at the same instant
# (inrush). When a test has multiple contactor_channel entries, they are
# closed one at a time with this delay between each.
CONTACTOR_STAGGER_S = 2.0

# Set the contactor line(s) for each test. A single channel string energizes
# one contactor; a list energizes each in turn, staggered by CONTACTOR_STAGGER_S.
TEST_DEFINITIONS = {
    'HV': {
        'label': 'HV Connectors',
        'contactor_channel': ['cDAQ2Mod1/line0', 'cDAQ2Mod1/line1'],
        'enabled': True,
        'influx_bucket_env': 'INFLUX_BUCKET_HV',
        'influx_token_env': 'INFLUX_TOKEN_HV',
        'max_current_a': 700,
        'flow_duration_min': 40,
        'bucket_tag': 'hv',
        'tc_channels': [
    ('cDAQ2Mod3/ai0',  'C01_SANTO'),
    ('cDAQ2Mod3/ai1',  'C02_COND'),
    ('cDAQ2Mod3/ai2',  'C03_SANTO'),
    ('cDAQ2Mod3/ai3',  'C04_COND'),
    ('cDAQ2Mod3/ai4',  'C05_SANTO'),
    ('cDAQ2Mod3/ai5',  'C06_COND'),
    ('cDAQ2Mod3/ai6',  'C07_SANTO'),
    ('cDAQ2Mod3/ai7',  'C08_COND'),
    ('cDAQ2Mod3/ai8',  'C09_SANTO'),
    ('cDAQ2Mod3/ai9',  'C10_COND'),
    ('cDAQ2Mod3/ai10', 'C11_SANTO'),
    ('cDAQ2Mod3/ai11', 'C12_COND'),
    ('cDAQ2Mod3/ai12', 'C13_SANTO'),
    ('cDAQ2Mod3/ai13', 'C14_COND'),
    ('cDAQ2Mod3/ai14', 'C15_SANTO'),
    ('cDAQ2Mod3/ai15', 'C16_COND'),
    ('cDAQ2Mod4/ai1',  'C18_COND'),
    ('cDAQ2Mod4/ai2',  'C19_SANTO'),
    ('cDAQ2Mod4/ai3',  'C20_SANTO'),
    ('cDAQ2Mod4/ai4',  'C21_SANTO'),
    ('cDAQ2Mod4/ai5',  'C22_SANTO'),
    ('cDAQ2Mod4/ai6',  'C23_COND'),
    ('cDAQ2Mod4/ai7',  'C24_SANTO'),
    ('cDAQ2Mod4/ai8',  'C25_SANTO'),
    ('cDAQ2Mod4/ai9',  'LUG-C26'),
    ('cDAQ2Mod4/ai10', 'TC27_CONTACTOR_AMB'),
    ('cDAQ2Mod4/ai11', 'C17_SANTO'),
    ('cDAQ2Mod4/ai12', 'TC29_CONTACTOR_2'),
    ('cDAQ2Mod4/ai13', 'TC30_CONTACTOR_3'),
    ('cDAQ2Mod4/ai15', 'TC31_TEST_AMB'),
    ('cDAQ2Mod4/ai14', 'TC28_CONTACTOR_1'),],
    },
    'FUSE': {
        'label': 'CCA Fuses',
        'contactor_channel': 'cDAQ2Mod1/line2',
        'enabled': True,
        'influx_bucket_env': 'INFLUX_BUCKET_FUSE',
        'influx_token_env': 'INFLUX_TOKEN_FUSE',
        'max_current_a': 215,
        'flow_duration_min': 35,
        'bucket_tag': 'fuse',
        'tc_channels': [],
    },
}

SEQUENCE_ORDER = list(TEST_DEFINITIONS.keys())

SHUNTS = [
    {'channel': 'cDAQ2Mod2/ai0', 'name': 'I_out',  'resistance_ohm': SHUNT_RESISTANCE_OHM},
]

TC_TYPE = 'T'
TC_INTERVAL = 10

# --- HV thermocouple over-temp safety trip ---------------------------------
# TCs are sampled every TC_INTERVAL (10s). If any channel is at/above its
# limit for OVER_TEMP_TRIP_READINGS consecutive samples, the sequence
# auto-pauses (PSU current to 0, contactors open).
# NOTE: placeholder values -- confirm against real hardware ratings before
# running unattended.
CONTACTOR_MAX_TEMP_C = 135.0          # TC28/29/30 (Contactor 1/2/3)
CONTACTOR_AMBIENT_MAX_TEMP_C = 75.0  # TC27 (Contactor Ambient)
SAMPLE_MAX_TEMP_C = 155.0            # everything else (C01-C26, TC31 Test Ambient)
OVER_TEMP_TRIP_READINGS = 2

CONTACTOR_TC_NAMES = {'TC28_CONT_1', 'TC29_CONT_2', 'TC30_CONT_3'}
CONTACTOR_AMBIENT_TC_NAMES = {'TC27_CONT_AMB'}

INFLUX_URL_ENV = 'INFLUX_URL'
INFLUX_ORG_ENV = 'INFLUX_ORG'
DEFAULT_INFLUX_ORG = 'tesse'

# ---------------------------------------------------------------------------
# Constants and helpers
# ---------------------------------------------------------------------------
PHASE_IDLE = 'idle'
PHASE_FLOW = 'flow'
PHASE_GRACE = 'grace'


def _other_test(name):
    idx = SEQUENCE_ORDER.index(name)
    return SEQUENCE_ORDER[(idx + 1) % len(SEQUENCE_ORDER)]


def _tc_temp_limit(name):
    if name in CONTACTOR_AMBIENT_TC_NAMES:
        return CONTACTOR_AMBIENT_MAX_TEMP_C
    if name in CONTACTOR_TC_NAMES:
        return CONTACTOR_MAX_TEMP_C
    return SAMPLE_MAX_TEMP_C


class SequenceState:
    def __init__(self):
        self.phase = PHASE_IDLE
        self.active_test = None
        self.manual_pause = False
        self.phase_started_at = None
        self.phase_deadline = None
        self.remaining_s = None
        self.cycle_counts = {name: 0 for name in TEST_DEFINITIONS}

    def to_dict(self):
        return {
            'phase': self.phase,
            'active_test': self.active_test,
            'manual_pause': self.manual_pause,
            'phase_started_at': self.phase_started_at,
            'phase_deadline': self.phase_deadline,
            'remaining_s': self.remaining_s,
            'cycle_counts': self.cycle_counts,
        }

    @classmethod
    def from_dict(cls, data):
        state = cls()
        state.phase = data.get('phase', PHASE_IDLE)
        state.active_test = data.get('active_test')
        state.manual_pause = data.get('manual_pause', False)
        state.phase_started_at = data.get('phase_started_at')
        state.phase_deadline = data.get('phase_deadline')
        state.remaining_s = data.get('remaining_s')
        counts = data.get('cycle_counts', {})
        state.cycle_counts.update({k: v for k, v in counts.items() if k in TEST_DEFINITIONS})
        return state

    def remaining_text(self):
        if self.phase_deadline is not None:
            remaining = max(0.0, self.phase_deadline - time.time())
        elif self.remaining_s is not None:
            remaining = self.remaining_s
        else:
            return 'n/a'
        return f'{remaining:.1f}s ({remaining / 60.0:.1f} min)'


class StateManager:
    def __init__(self, state_file):
        self.state_file = state_file
        self.lock = threading.RLock()
        self.state = SequenceState()

    def load(self):
        if os.path.isfile(self.state_file):
            try:
                with open(self.state_file, 'r', encoding='utf-8') as f:
                    raw = json.load(f)
                self.state = SequenceState.from_dict(raw.get('sequence', {}))
            except Exception:
                print('Warning: failed to load state file, starting fresh.')
                self.state = SequenceState()
        else:
            self.state = SequenceState()

    def save(self):
        with self.lock:
            output = {'sequence': self.state.to_dict(), 'saved_at': time.time()}
            with open(self.state_file, 'w', encoding='utf-8') as f:
                json.dump(output, f, indent=2)


# ---------------------------------------------------------------------------
# Instrument classes
# ---------------------------------------------------------------------------
class PSU:
    def __init__(self, visa_address, voltage_limit, simulate):
        self.simulate = simulate
        self._voltage_limit = voltage_limit
        self._current_sp = 0.0
        self._psu = None

        if simulate:
            print(f'  [sim] PSU initialised ({visa_address}) — output OFF, 0 A')
            return

        import pyvisa
        rm = pyvisa.ResourceManager()
        psu = rm.open_resource(visa_address)
        psu.timeout = 5000
        psu.write_termination = '\n'
        psu.read_termination = '\n'
        idn = psu.query('*IDN?').strip()
        print(f'  PSU connected: {idn}')
        self._psu = psu
        psu.write(f'SOUR:VOLT {voltage_limit:.3f}')
        psu.write('SOUR:CURR 0')
        psu.write('OUTP OFF')

    def output_on(self):
        if self.simulate:
            print('  [sim] PSU output ON')
            return
        self._psu.write('OUTP ON')

    def set_current(self, amps):
        self._current_sp = float(amps)
        if self.simulate:
            print(f'  [sim] PSU → {amps:.3f} A')
            return
        self._psu.write(f'SOUR:CURR {amps:.3f}')

    def measure(self):
        if self.simulate:
            return (self._current_sp + 0.1, 3.0)
        current_resp = self._psu.query('MEAS:CURR?').strip()
        voltage_resp = self._psu.query('MEAS:VOLT?').strip()
        try:
            i = float(current_resp)
            v = float(voltage_resp)
        except ValueError:
            raise ValueError(
                f"Unexpected PSU response during measurement: CURR?='{current_resp}', VOLT?='{voltage_resp}'"
            )
        return i, v

    def shutdown(self, reason=''):
        if self.simulate:
            print(f'  [sim] PSU shutdown{f" — {reason}" if reason else ""}')
            self._current_sp = 0.0
            return
        if self._psu is None:
            return
        for attempt in range(3):
            try:
                self._psu.write('SOUR:CURR 0')
                self._psu.write('OUTP OFF')
                self._current_sp = 0.0
                print(f'  PSU output OFF{f" — {reason}" if reason else ""}')
                return
            except Exception as exc:
                print(f'  [PSU] shutdown attempt {attempt + 1}/3 failed: {exc}')
                time.sleep(0.5)
        print('  !! PSU SHUTDOWN FAILED — MANUALLY VERIFY OUTPUT IS OFF !!')

    def close(self):
        self.shutdown()
        if self._psu is not None:
            self._psu.close()
            self._psu = None


class Contactor:
    def __init__(self, channel, simulate, label):
        self.channel = channel
        self.simulate = simulate
        self.label = label
        self._task = None

        if simulate:
            print(f'  [sim] Contactor {label} ({channel}) initialised — OPEN')
            return

        import nidaqmx
        self._task = nidaqmx.Task()
        self._task.do_channels.add_do_chan(channel)
        self._task.start()
        self._write(False)

    def energize(self):
        if self.simulate:
            print(f'  [sim] Contactor {self.label} CLOSED')
            return
        self._write(True)

    def de_energize(self):
        if self.simulate:
            print(f'  [sim] Contactor {self.label} OPENED')
            return
        self._write(False)

    def _write(self, state):
        self._task.write(state)

    def close(self):
        if self._task is not None:
            self._write(False)
            self._task.close()
            self._task = None


class ContactorGroup:
    """Drives one or more Contactor instances as a unit. When more than one
    contactor is present, energize() closes them one at a time with
    stagger_s between each so contactors sharing a supply don't inrush
    simultaneously. de_energize() opens them all immediately."""

    def __init__(self, contactors, stagger_s=0.0):
        self.contactors = contactors
        self.stagger_s = stagger_s

    def energize(self):
        for i, contactor in enumerate(self.contactors):
            if i > 0 and self.stagger_s > 0:
                time.sleep(self.stagger_s)
            contactor.energize()

    def de_energize(self):
        for contactor in self.contactors:
            contactor.de_energize()

    def close(self):
        for contactor in self.contactors:
            contactor.close()


class ShuntReader:
    def __init__(self, shunts, max_voltage, simulate):
        self.shunts = shunts
        self.simulate = simulate
        self._t0 = time.time()
        self._task = None

        if simulate:
            return

        import nidaqmx
        from nidaqmx.constants import TerminalConfiguration
        self._task = nidaqmx.Task()
        for s in shunts:
            self._task.ai_channels.add_ai_voltage_chan(
                s['channel'],
                terminal_config=TerminalConfiguration.DIFF,
                min_val=-max_voltage,
                max_val=max_voltage,
            )
        self._task.start()

    def read_currents(self):
        if self.simulate:
            elapsed = time.time() - self._t0
            nominal = min(15.0, 15.0 * elapsed / 10.0)
            return [nominal + 0.1 * (i - 1) for i in range(len(self.shunts))]

        raw = self._task.read()
        voltages = raw if isinstance(raw, list) else [raw]
        return [v / s['resistance_ohm'] for v, s in zip(voltages, self.shunts)]

    def close(self):
        if self._task is not None:
            self._task.close()
            self._task = None


class TCReader:
    def __init__(self, channels, tc_type, simulate):
        self.channels = channels
        self.tc_type = tc_type
        self.simulate = simulate
        self._t0 = time.time()
        self._task = None

        if simulate:
            return

        import nidaqmx
        from nidaqmx.constants import ThermocoupleType, TemperatureUnits
        tmap = {
            'J': ThermocoupleType.J,
            'K': ThermocoupleType.K,
            'T': ThermocoupleType.T,
            'E': ThermocoupleType.E,
            'N': ThermocoupleType.N,
        }
        self._task = nidaqmx.Task()
        for channel, _ in channels:
            self._task.ai_channels.add_ai_thrmcpl_chan(
                channel,
                thermocouple_type=tmap[tc_type],
                units=TemperatureUnits.DEG_C,
            )
        self._task.start()

    def read(self):
        if self.simulate:
            elapsed = time.time() - self._t0
            return [25.0 + 2.0 * i + 0.5 * (math.sin(elapsed) if i % 2 == 0 else math.cos(elapsed))
                    for i in range(len(self.channels))]

        raw = self._task.read()
        return raw if isinstance(raw, list) else [raw]

    def names(self):
        return [name for _, name in self.channels]

    def close(self):
        if self._task is not None:
            self._task.close()
            self._task = None


class InfluxWriter:
    def __init__(self, url, org, bucket, token, test_tag):
        import influxdb_client
        from influxdb_client.client.write_api import SYNCHRONOUS
        self._ic = influxdb_client
        self._client = influxdb_client.InfluxDBClient(url=url, token=token, org=org)
        self._writer = self._client.write_api(write_options=SYNCHRONOUS)
        self._bucket = bucket
        self._org = org
        self._test_tag = test_tag

    def write_currents(self, currents, shunts, timestamp):
        ts_ms = round(timestamp * 1000.0)
        points = []
        for current, s in zip(currents, shunts):
            points.append(
                self._ic.Point('current')
                .tag('test', self._test_tag)
                .tag('shunt', s['name'])
                .field('amps', float(current))
                .time(ts_ms, self._ic.WritePrecision.MS)
            )
        self._writer.write(bucket=self._bucket, org=self._org, record=points)

    def write_tcs(self, temperatures, names, timestamp):
        ts_ms = round(timestamp * 1000.0)
        points = []
        for name, temp in zip(names, temperatures):
            points.append(
                self._ic.Point('temperature')
                .tag('test', self._test_tag)
                .tag('channel', name)
                .field('deg_c', float(temp))
                .time(ts_ms, self._ic.WritePrecision.MS)
            )
        self._writer.write(bucket=self._bucket, org=self._org, record=points)

    def close(self):
        self._client.close()


class NullTCWriter:
    def __init__(self, test_tag):
        self._test_tag = test_tag

    def write_tcs(self, temperatures, names, timestamp):
        values = ', '.join(f'{name}={temp:.2f}C' for name, temp in zip(names, temperatures))
        print(f'  [tc log {self._test_tag}] {values}')

    def close(self):
        pass


class CsvLogger:
    def __init__(self, test_name, test_label):
        self.test_name = test_name
        self.test_label = test_label
        self.filename = LOG_FILE_CSV_TEMPLATE.format(test=test_name.lower())
        self._ensure_file()

    def _ensure_file(self):
        if not os.path.isfile(self.filename):
            with open(self.filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    'timestamp',
                    'cycle',
                    'test',
                    'phase',
                    'current_owner',
                    'psu_current_a',
                    'psu_voltage_v',
                    'shunt_name',
                    'shunt_amps',
                ])

    def write_currents(self, currents, shunts, timestamp, cycle, phase, current_owner, psu_current, psu_voltage):
        iso_ts = datetime.fromtimestamp(timestamp).isoformat(sep=' ')
        rows = []
        for current, s in zip(currents, shunts):
            rows.append([
                iso_ts,
                cycle,
                self.test_name,
                phase,
                current_owner or '',
                round(psu_current, 6) if psu_current is not None else '',
                round(psu_voltage, 6) if psu_voltage is not None else '',
                s['name'],
                round(current, 6),
            ])
        with open(self.filename, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(rows)

    def close(self):
        pass


class ExcelLogger:
    def __init__(self, test_name, test_label):
        self.test_name = test_name
        self.test_label = test_label
        self.filename = LOG_FILE_XLSX
        self.sheet_name = test_name.upper()
        self._workbook = None
        self._worksheet = None
        self._init_workbook()

    def _init_workbook(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self._workbook = None
            return

        if os.path.isfile(self.filename):
            try:
                self._workbook = load_workbook(self.filename)
            except Exception:
                self._workbook = Workbook()
        else:
            self._workbook = Workbook()

        if self.sheet_name in self._workbook.sheetnames:
            self._worksheet = self._workbook[self.sheet_name]
        else:
            self._worksheet = self._workbook.create_sheet(self.sheet_name)
            self._worksheet.append([
                'timestamp',
                'cycle',
                'test',
                'phase',
                'current_owner',
                'psu_current_a',
                'psu_voltage_v',
                'shunt_name',
                'shunt_amps',
            ])

    def write_currents(self, currents, shunts, timestamp, cycle, phase, current_owner, psu_current, psu_voltage):
        if self._workbook is None or self._worksheet is None:
            return
        iso_ts = datetime.fromtimestamp(timestamp).isoformat(sep=' ')
        for current, s in zip(currents, shunts):
            self._worksheet.append([
                iso_ts,
                cycle,
                self.test_name,
                phase,
                current_owner or '',
                round(psu_current, 6) if psu_current is not None else '',
                round(psu_voltage, 6) if psu_voltage is not None else '',
                s['name'],
                round(current, 6),
            ])

    def close(self):
        if self._workbook is not None:
            try:
                self._workbook.save(self.filename)
            except Exception as exc:
                print(f'  [ExcelLogger] failed to save {self.filename}: {exc}')


class NullWriter:
    def __init__(self, test_tag):
        self.test_tag = test_tag

    def write_currents(self, currents, shunts, timestamp, cycle=None, phase=None, current_owner=None):
        values = '  '.join(f"{s['name']}={c:.2f}A" for c, s in zip(currents, shunts))
        print(f'  [log {self.test_tag}] {values}')

    def close(self):
        pass


# ---------------------------------------------------------------------------
# Runtime manager
# ---------------------------------------------------------------------------
class DualTestRuntime:
    def __init__(self):
        self._stop_event = threading.Event()
        self._state_manager = StateManager(STATE_FILE)
        self._state_manager.load()
        self._seq = self._state_manager.state
        self._tc_readers = {}
        self._tc_writers = {}
        self._contactors = {}
        self._tc_over_counts = {}
        self._psu = PSU(PSU_VISA_ADDRESS, PSU_VOLTAGE_LIMIT, SIMULATE)
        self._last_state_save = time.time()
        self._last_tc_log = time.time()
        self._load_tc_writers()
        self._create_contactors()
        self._reconcile_electrical_state()
        self._start_if_needed()

    def _load_tc_writers(self):
        url = os.environ[INFLUX_URL_ENV]
        org = os.environ.get(INFLUX_ORG_ENV, DEFAULT_INFLUX_ORG)
        for name, cfg in TEST_DEFINITIONS.items():
            bucket = os.environ.get(cfg['influx_bucket_env'])
            if not bucket:
                raise RuntimeError(f'Missing env var {cfg["influx_bucket_env"]} for {name}')
            token = os.environ.get(cfg['influx_token_env'])
            if not token:
                raise RuntimeError(f'Missing env var {cfg["influx_token_env"]} for {name}')
            if cfg.get('tc_channels'):
                if SIMULATE:
                    self._tc_writers[name] = NullTCWriter(cfg['bucket_tag'])
                else:
                    self._tc_writers[name] = InfluxWriter(url, org, bucket, token, cfg['bucket_tag'])
                self._tc_readers[name] = TCReader(cfg['tc_channels'], TC_TYPE, SIMULATE)

    def _create_contactors(self):
        for name, cfg in TEST_DEFINITIONS.items():
            channels = cfg['contactor_channel']
            if isinstance(channels, (list, tuple)):
                contactors = [
                    Contactor(channel, SIMULATE, f"{cfg['label']} #{i + 1}")
                    for i, channel in enumerate(channels)
                ]
            else:
                contactors = [Contactor(channels, SIMULATE, cfg['label'])]
            self._contactors[name] = ContactorGroup(contactors, stagger_s=CONTACTOR_STAGGER_S)

    def _reconcile_electrical_state(self):
        seq = self._seq
        if seq.phase == PHASE_IDLE or seq.manual_pause or seq.active_test is None:
            return
        cfg = TEST_DEFINITIONS[seq.active_test]
        if seq.phase == PHASE_FLOW:
            self._contactors[seq.active_test].energize()
            self._psu.output_on()
            self._psu.set_current(cfg['max_current_a'])
        elif seq.phase == PHASE_GRACE:
            self._contactors[seq.active_test].energize()
            self._psu.shutdown('resumed after restart, still in grace period')

    def _start_if_needed(self):
        if self._seq.phase == PHASE_IDLE and AUTO_START:
            self.start_sequence()

    def _enter_flow(self, name, now):
        cfg = TEST_DEFINITIONS[name]
        previous = self._seq.active_test
        if previous is not None and previous != name:
            self._contactors[previous].de_energize()
        self._contactors[name].energize()
        self._psu.output_on()
        self._psu.set_current(cfg['max_current_a'])
        self._seq.phase = PHASE_FLOW
        self._seq.active_test = name
        self._seq.cycle_counts[name] += 1
        self._seq.phase_started_at = now
        self._seq.remaining_s = cfg['flow_duration_min'] * 60
        self._seq.phase_deadline = now + self._seq.remaining_s
        self._seq.manual_pause = False
        self._append_history('flow_start', name)
        self._state_manager.save()
        print(f'{cfg["label"]} flowing at {cfg["max_current_a"]} A (cycle {self._seq.cycle_counts[name]})')

    def _enter_grace(self, now):
        finishing = self._seq.active_test
        upcoming = _other_test(finishing)
        self._psu.shutdown(f'grace period before {TEST_DEFINITIONS[upcoming]["label"]}')
        self._seq.phase = PHASE_GRACE
        self._seq.phase_started_at = now
        self._seq.remaining_s = GRACE_PERIOD_S
        self._seq.phase_deadline = now + GRACE_PERIOD_S
        self._seq.manual_pause = False
        self._append_history('grace_start', finishing)
        self._state_manager.save()

    def _enter_idle(self, now):
        self._psu.shutdown('sequence stopped')
        if self._seq.active_test is not None:
            self._contactors[self._seq.active_test].de_energize()
        self._seq.phase = PHASE_IDLE
        self._seq.active_test = None
        self._seq.phase_started_at = now
        self._seq.phase_deadline = None
        self._seq.remaining_s = None
        self._seq.manual_pause = False
        self._append_history('idle', None)
        self._state_manager.save()

    def _append_history(self, event, test_name):
        entry = {
            'timestamp': time.time(),
            'test': test_name,
            'event': event,
            'phase': self._seq.phase,
            'cycle': self._seq.cycle_counts.get(test_name) if test_name else None,
            'manual_pause': self._seq.manual_pause,
        }
        with open(HISTORY_FILE, 'a', encoding='utf-8') as f:
            f.write(json.dumps(entry) + '\n')

    def _update_sequence(self, now):
        seq = self._seq
        if seq.phase == PHASE_IDLE or seq.manual_pause or seq.phase_deadline is None:
            return
        if now < seq.phase_deadline:
            return
        if seq.phase == PHASE_FLOW:
            self._enter_grace(now)
        elif seq.phase == PHASE_GRACE:
            self._enter_flow(_other_test(seq.active_test), now)

    def _update_states(self):
        now = time.time()
        self._update_sequence(now)
        if time.time() - self._last_state_save > STATE_SAVE_INTERVAL_S:
            self._state_manager.save()
            self._last_state_save = time.time()

    def _maybe_log_tcs(self, timestamp):
        if 'HV' not in self._tc_readers or 'HV' not in self._tc_writers:
            return
        try:
            temperatures = self._tc_readers['HV'].read()
            names = self._tc_readers['HV'].names()
            self._tc_writers['HV'].write_tcs(temperatures, names, timestamp)
            self._check_tc_safety(names, temperatures)
        except Exception as exc:
            print(f'  [tc logger error] {type(exc).__name__}: {exc}')

    def _check_tc_safety(self, names, temperatures):
        tripped = []
        for name, temp in zip(names, temperatures):
            limit = _tc_temp_limit(name)
            if temp >= limit:
                count = self._tc_over_counts.get(name, 0) + 1
                self._tc_over_counts[name] = count
                if count >= OVER_TEMP_TRIP_READINGS:
                    tripped.append((name, temp, limit))
            else:
                self._tc_over_counts[name] = 0

        if not tripped:
            return
        details = ', '.join(f'{name}={temp:.1f}C (limit {limit:.1f}C)' for name, temp, limit in tripped)
        if self._seq.phase == PHASE_IDLE or self._seq.manual_pause:
            return
        print(f'\n!! OVER-TEMP TRIP: {details} — auto-pausing sequence !!')
        self._append_history('overtemp_trip', self._seq.active_test)
        self.pause_sequence()

    def _write_status(self):
        seq = self._seq
        label = TEST_DEFINITIONS[seq.active_test]['label'] if seq.active_test else 'none'
        state_text = 'paused' if seq.manual_pause else 'running'
        counts = ', '.join(f'{name}={seq.cycle_counts[name]}' for name in TEST_DEFINITIONS)
        print('\nCurrent runtime status:')
        print(f'  phase {seq.phase:<5} | active {label:<14} | {state_text:<7} | remain {seq.remaining_text()}')
        print(f'  cycle counts: {counts}')

    def start_sequence(self, first_test=None):
        seq = self._seq
        if seq.phase != PHASE_IDLE:
            print('Sequence is already running (use pause/resume/stop).')
            return
        name = (first_test or SEQUENCE_ORDER[0]).upper()
        if name not in TEST_DEFINITIONS:
            print(f'Unknown test {name}. Choose one of: {", ".join(TEST_DEFINITIONS)}')
            return
        self._enter_flow(name, time.time())
        print(f'Started sequence, beginning with {TEST_DEFINITIONS[name]["label"]}')

    def pause_sequence(self):
        seq = self._seq
        if seq.phase == PHASE_IDLE:
            print('Sequence is idle.')
            return
        if seq.manual_pause:
            print('Sequence is already paused.')
            return
        seq.manual_pause = True
        if seq.phase_deadline is not None:
            seq.remaining_s = max(0.0, seq.phase_deadline - time.time())
            seq.phase_deadline = None
        if seq.phase == PHASE_FLOW:
            self._psu.shutdown('manual pause')
            self._contactors[seq.active_test].de_energize()
        self._state_manager.save()
        print('Sequence paused.')

    def resume_sequence(self):
        seq = self._seq
        if not seq.manual_pause:
            print('Sequence is not paused.')
            return
        seq.manual_pause = False
        if seq.remaining_s is None:
            seq.remaining_s = 0.0
        seq.phase_deadline = time.time() + seq.remaining_s
        if seq.phase == PHASE_FLOW:
            cfg = TEST_DEFINITIONS[seq.active_test]
            self._contactors[seq.active_test].energize()
            self._psu.output_on()
            self._psu.set_current(cfg['max_current_a'])
        self._state_manager.save()
        print('Sequence resumed.')

    def stop_sequence(self):
        if self._seq.phase == PHASE_IDLE:
            print('Sequence is already idle.')
            return
        self._enter_idle(time.time())
        print('Sequence stopped.')

    def skip_phase(self):
        seq = self._seq
        if seq.phase == PHASE_IDLE:
            print('Sequence is idle; nothing to skip.')
            return
        if seq.manual_pause:
            print('Sequence is paused; resume before skipping.')
            return
        now = time.time()
        if seq.phase == PHASE_FLOW:
            finishing = TEST_DEFINITIONS[seq.active_test]['label']
            self._enter_grace(now)
            print(f'Skipped remaining flow time for {finishing}; entering grace period.')
        elif seq.phase == PHASE_GRACE:
            upcoming = _other_test(seq.active_test)
            self._enter_flow(upcoming, now)
            print(f'Skipped grace period; {TEST_DEFINITIONS[upcoming]["label"]} now flowing.')

    def dump_state(self):
        self._state_manager.save()
        print(f'State saved to {STATE_FILE}')

    def close(self):
        self._stop_event.set()
        self._psu.shutdown('application exit')
        for contactor in self._contactors.values():
            contactor.close()
        for writer in self._tc_writers.values():
            writer.close()
        for reader in self._tc_readers.values():
            reader.close()
        self._state_manager.save()
        print('Shutdown complete.')

    def _publish_live_readings(self, timestamp, psu_current, psu_voltage):
        payload = {
            'timestamp': timestamp,
            'psu_current_a': psu_current,
            'psu_voltage_v': psu_voltage,
        }
        with open(LIVE_READINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(payload, f)

    def readings_publish_loop(self):
        while not self._stop_event.is_set():
            timestamp = time.time()
            try:
                psu_current, psu_voltage = self._psu.measure()
                self._publish_live_readings(timestamp, psu_current, psu_voltage)
            except Exception as exc:
                print(f'  [readings publish error] {type(exc).__name__}: {exc}')
            if (timestamp - self._last_tc_log) >= TC_INTERVAL:
                self._maybe_log_tcs(timestamp)
                self._last_tc_log = timestamp
            self._stop_event.wait(LOG_INTERVAL)

    def command_loop(self):
        self._write_status()
        self.print_help()
        while not self._stop_event.is_set():
            try:
                command = input('> ').strip().lower()
            except EOFError:
                break
            if not command:
                continue
            self.handle_command(command)
            if self._stop_event.is_set():
                break

    def handle_command(self, command):
        tokens = command.split()
        if not tokens:
            return
        cmd = tokens[0]
        args = tokens[1:]
        if cmd in ('help', '?'):
            self.print_help()
        elif cmd == 'status':
            self._write_status()
        elif cmd == 'start':
            self.start_sequence(args[0] if args else None)
        elif cmd == 'pause':
            self.pause_sequence()
        elif cmd == 'resume':
            self.resume_sequence()
        elif cmd == 'stop':
            self.stop_sequence()
        elif cmd == 'skip':
            self.skip_phase()
        elif cmd == 'dump':
            self.dump_state()
        elif cmd in ('quit', 'exit', 'q'):
            self._stop_event.set()
        else:
            print(f'Unknown command: {command}')
            self.print_help()

    def print_help(self):
        print('\nCommands:')
        print('  status               - show current runtime state')
        print('  start [HV|FUSE]      - begin the alternating sequence (defaults to HV first)')
        print('  pause                - pause the sequence, current off')
        print('  resume               - resume a paused sequence')
        print('  stop                 - stop the sequence and reset to idle')
        print('  skip                 - end the current phase early and advance')
        print('  dump                 - save current state to disk')
        print('  quit / exit / q      - cleanly shutdown and exit')

    def run(self):
        signal.signal(signal.SIGINT, self._signal_handler)
        signal.signal(signal.SIGTERM, self._signal_handler)
        readings_thread = threading.Thread(target=self.readings_publish_loop, daemon=True)
        readings_thread.start()
        try:
            while not self._stop_event.is_set():
                self._update_states()
                time.sleep(0.2)
        except KeyboardInterrupt:
            self._stop_event.set()
        finally:
            self.close()

    def _signal_handler(self, signum, frame):
        print(f'\nReceived signal {signum}, shutting down...')
        self._stop_event.set()


def main():
    from dotenv import load_dotenv
    load_dotenv()

    runtime = DualTestRuntime()
    print('Dual test runtime loaded. Type "help" for commands.')
    command_thread = threading.Thread(target=runtime.command_loop, daemon=True)
    command_thread.start()
    runtime.run()


if __name__ == '__main__':
    main()
